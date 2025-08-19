import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# ---------- File Names ----------
INPUT_FILE = "marks.xlsx"       # input file (your marks sheet)
OUTPUT_FILE = "marks_report.xlsx"  # output report file

# ---------- Required Columns ----------
REQUIRED_COLS = ["Name", "Math", "Science", "English"]

# ---------- Grading Function ----------
def get_grade(pct: float) -> str:
    if pct >= 90:
        return "A+"
    elif pct >= 75:
        return "A"
    elif pct >= 60:
        return "B"
    else:
        return "C"

# ---------- Main Program ----------
def main():
    # Step 1: Load Excel
    try:
        df = pd.read_excel(INPUT_FILE, engine="openpyxl")
    except FileNotFoundError:
        print(f"❌ Could not find '{INPUT_FILE}'. Make sure the file exists.")
        sys.exit(1)
    except Exception as e:
        print("❌ Failed to read Excel:", e)
        sys.exit(1)

    # Step 2: Validate Columns
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns: {missing}")
        print(f"   Found columns: {list(df.columns)}")
        sys.exit(1)

    # Step 3: Convert Marks to Numbers
    for col in ["Math", "Science", "English"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Step 4: Calculate Total, Percentage, Grade
    df["Total"] = df["Math"] + df["Science"] + df["English"]
    df["Percentage"] = (df["Total"] / 3).round(2)
    df["Grade"] = df["Percentage"].apply(get_grade)

    # Step 5: Save Report to Excel
    df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    # ---------- Step 6: Apply Color Coding ----------
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")     # Fail
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid") # Average
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Good
    blue_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")   # Excellent

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    subject_cols = {"Math": 2, "Science": 3, "English": 4}

    for row in range(2, ws.max_row + 1):  # start at row 2 (skip headers)
        for subject, col in subject_cols.items():
            cell = ws.cell(row=row, column=col)
            mark = cell.value

            if mark < 40:
                cell.fill = red_fill
            elif mark < 60:
                cell.fill = yellow_fill
            elif mark < 80:
                cell.fill = green_fill
            else:
                cell.fill = blue_fill

    # ---------- Step 7: Add Neat Color Legend (Right Side) ----------
    legend_col = ws.max_column + 2   # place legend 2 columns away from data
    legend_row = 2                   # start near top

    # Define a thin border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.cell(row=legend_row, column=legend_col, value="Color Legend").border = thin_border

    ws.cell(row=legend_row + 1, column=legend_col, value="< 40 = Fail").fill = red_fill
    ws.cell(row=legend_row + 1, column=legend_col).border = thin_border

    ws.cell(row=legend_row + 2, column=legend_col, value="40–59 = Average").fill = yellow_fill
    ws.cell(row=legend_row + 2, column=legend_col).border = thin_border

    ws.cell(row=legend_row + 3, column=legend_col, value="60–79 = Good").fill = green_fill
    ws.cell(row=legend_row + 3, column=legend_col).border = thin_border

    ws.cell(row=legend_row + 4, column=legend_col, value="80+ = Excellent").fill = blue_fill
    ws.cell(row=legend_row + 4, column=legend_col).border = thin_border

    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"✅ Report generated successfully with neat color coding and legend: {OUTPUT_FILE}")

# ---------- Run Program ----------
if __name__ == "__main__":
    main()
